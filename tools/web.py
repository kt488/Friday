import requests
from bs4 import BeautifulSoup
import urllib.parse
import time
import os
import csv
import json
import io

class WebTools:
    @staticmethod
    def parse_file(file_path, output_format="txt"):
        """
        Parses a file and returns its contents in the requested format.

        Supported file types: .txt, .csv, .json, .xml, .pdf, .xlsx, .xls, .html
        Supported output formats: txt, csv, json

        Examples:
          parse_file("workspace/data.xlsx", "csv")   -> Excel data as CSV text
          parse_file("temp/report.pdf", "txt")   -> PDF text content
          parse_file("temp/data.json", "json")   -> JSON data (pretty-printed)
        """
        if not os.path.exists(file_path):
            return f"Error: File not found at '{file_path}'"

        ext = os.path.splitext(file_path)[1].lower()
        raw_data = None

        try:
            # --- TXT ---
            if ext == '.txt':
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    raw_data = f.read()

            # --- CSV ---
            elif ext == '.csv':
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    reader = csv.DictReader(f)
                    raw_data = list(reader)

            # --- JSON ---
            elif ext == '.json':
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    raw_data = json.load(f)

            # --- XML ---
            elif ext == '.xml':
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    soup = BeautifulSoup(f.read(), 'xml')
                    raw_data = soup.prettify()

            # --- PDF ---
            elif ext == '.pdf':
                try:
                    import PyPDF2
                except ImportError:
                    return "Error: PyPDF2 is not installed (pip install PyPDF2)"
                with open(file_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    pages = []
                    for page in reader.pages:
                        t = page.extract_text()
                        if t:
                            pages.append(t)
                    raw_data = "\n--- Page ---\n".join(pages)

            # --- Excel ---
            elif ext in ('.xlsx', '.xls'):
                try:
                    import openpyxl
                except ImportError:
                    return "Error: openpyxl is not installed (pip install openpyxl)"
                import re
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets = {}
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append([str(c) if c is not None else "" for c in row])
                    sheets[sheet_name] = rows
                wb.close()
                raw_data = sheets

            # --- HTML ---
            elif ext in ('.html', '.htm'):
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    soup = BeautifulSoup(f.read(), 'html.parser')
                for tag in soup(["script", "style", "nav", "footer", "header"]):
                    tag.decompose()
                raw_data = soup.get_text(separator="\n", strip=True)

            else:
                return f"Error: Unsupported file type '{ext}'. Supported: .txt, .csv, .json, .xml, .pdf, .xlsx, .xls, .html"

        except Exception as e:
            return f"Error parsing file: {e}"

        # --- Format output ---
        return WebTools._format_output(raw_data, output_format)

    @staticmethod
    def _format_output(data, output_format):
        """Internal helper: converts parsed data to the requested format."""
        fmt = output_format.lower()

        if isinstance(data, str):
            if fmt == "json":
                return json.dumps({"content": data}, indent=2)
            return data  # txt/csv on a string just returns it

        if fmt == "json":
            return json.dumps(data, indent=2, default=str, ensure_ascii=False)

        if fmt == "csv":
            if isinstance(data, list) and data:
                buf = io.StringIO()
                if isinstance(data[0], dict):
                    w = csv.DictWriter(buf, fieldnames=data[0].keys())
                    w.writeheader()
                    w.writerows(data)
                elif isinstance(data[0], (list, tuple)):
                    w = csv.writer(buf)
                    w.writerows(data)
                else:
                    w = csv.writer(buf)
                    for item in data:
                        w.writerow([item])
                return buf.getvalue()
            return str(data)

        # fmt == "txt" (default)
        if isinstance(data, dict):
            lines = []
            for key, value in data.items():
                if isinstance(value, list):
                    lines.append(f"\n--- {key} ---")
                    for row in value:
                        lines.append("\t".join(str(c) for c in row))
                else:
                    lines.append(f"{key}: {value}")
            return "\n".join(lines)
        if isinstance(data, list):
            return "\n".join(
                json.dumps(item, default=str) if isinstance(item, dict) else str(item)
                for item in data
            )
        return str(data)

    @staticmethod
    def web_search(query, num_results=5):
        """Searches the web using DuckDuckGo and returns a list of result titles and URLs."""
        try:
            # We use the DuckDuckGo HTML version which is easier to scrape
            url = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote(query)}"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Referer': 'https://duckduckgo.com/'
            }
            
            response = requests.get(url, headers=headers, timeout=15)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            results = []
            
            # Find all result containers
            result_containers = soup.find_all('div', class_='result')
            
            for i, container in enumerate(result_containers):
                if i >= num_results:
                    break
                    
                title_tag = container.find('a', class_='result__a')
                snippet_tag = container.find('a', class_='result__snippet')
                
                if title_tag:
                    title = title_tag.get_text(strip=True)
                    link = title_tag.get('href')
                    
                    # Clean up the link if it's a redirect
                    if link and link.startswith('//'):
                        link = 'https:' + link
                    
                    if link and '/l/?' in link and 'uddg=' in link:
                        try:
                            import re
                            match = re.search(r'uddg=([^&]+)', link)
                            if match:
                                link = urllib.parse.unquote(match.group(1))
                        except Exception:
                            pass
                        
                    snippet = snippet_tag.get_text(strip=True) if snippet_tag else "No description available."
                    
                    results.append(f"Title: {title}\nURL: {link}\nDescription: {snippet}\n")
            
            if not results:
                return "No results found. The search engine might be blocking the request or the query returned nothing."
                
            return "\n".join(results)
        except Exception as e:
            return f"Error during web search: {e}"

    @staticmethod
    def scrape_content(url):
        """Scrapes the main text content from a given URL."""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            response = requests.get(url, headers=headers, timeout=15)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()

            # Get text
            text = soup.get_text()

            # Break into lines and remove leading and trailing whitespace
            lines = (line.strip() for line in text.splitlines())
            # Break multi-headlines into a line each
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            # Drop blank lines
            text = '\n'.join(chunk for chunk in chunks if chunk)

            # Limit text length to avoid context overflow
            return text[:5000] + "..." if len(text) > 5000 else text
        except Exception as e:
            return f"Error during scraping: {e}"

def get_tool_map():
    """Returns a map of tool names to their functions."""
    return {
        "web_search": WebTools.web_search,
        "scrape_content": WebTools.scrape_content,
        "parse_file": WebTools.parse_file
    }
